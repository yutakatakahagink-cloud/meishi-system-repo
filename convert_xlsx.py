# -*- coding: utf-8 -*-
"""
名刺ﾃﾞｰﾀ.xlsx を Webアプリ用 JSON (meishi-app/public/data/meishi-records.json) に変換する。
Excel を更新したら、このスクリプトを実行して JSON を作り直す。

  python convert_xlsx.py
"""
import openpyxl, json, os

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, '名刺ﾃﾞｰﾀ.xlsx')
OUTDIR = os.path.join(BASE, 'meishi-app', 'public', 'data')
OUT = os.path.join(OUTDIR, 'meishi-records.json')

KEYMAP = {
    '番号': 'no', '氏名': 'name', '会社・団体名': 'company',
    '所属1': 'aff1', '所属2': 'aff2', '所属3': 'aff3',
    '役職': 'title', '資格': 'qual', '携帯': 'mobile', 'メール': 'email',
    '郵便番号': 'postal', '住所': 'address', 'TEL': 'tel', 'FAX': 'fax',
    'URL': 'url', '備考': 'note', '区分': 'category',
}


def main():
    os.makedirs(OUTDIR, exist_ok=True)
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    recs = []
    for r in range(2, ws.max_row + 1):
        rowvals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in rowvals):
            continue
        rec = {}
        for h, v in zip(headers, rowvals):
            key = KEYMAP.get(str(h).strip() if h is not None else '', None)
            if key is None:
                continue
            rec[key] = '' if v is None else str(v).strip()
        recs.append(rec)
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(recs, f, ensure_ascii=False, indent=1)
    print(f"{len(recs)} 件を書き出しました -> {OUT}")


if __name__ == '__main__':
    main()
